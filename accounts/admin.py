from django.contrib import admin
from django.contrib.auth.models import Group
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path, reverse
from django.contrib import messages
from django.shortcuts import render
import csv
import xlsxwriter
import openpyxl
from io import BytesIO, TextIOWrapper
from unfold.admin import ModelAdmin
from unfold.decorators import action as unfold_action   # ← QO'SHILDI
from .models import Student

# Groups modelini admin paneldan yashirish
admin.site.unregister(Group)


@admin.register(Student)
class StudentAdmin(ModelAdmin):
    change_list_template = 'admin/accounts/student/change_list.html'
    list_display = ('hemis_id', 'full_name', 'faculty', 'year', 'birth_date', 'is_active', 'date_joined')
    list_filter = ('is_active', 'faculty', 'year')
    search_fields = ('hemis_id', 'full_name', 'faculty')
    ordering = ('-date_joined',)
    readonly_fields = ('date_joined', 'last_login')

    # actions_list — Unfold tugmalari (unfold_action kerak)
    # actions      — oddiy Django dropdown (admin.action kerak)
    actions_list = ['export_to_excel', 'export_to_csv']
    actions = ['export_to_excel', 'export_to_csv']

    fieldsets = (
        (None, {'fields': ('hemis_id',)}),
        ('Personal info', {'fields': ('full_name', 'faculty', 'year', 'birth_date')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'user_permissions')}),
        ('Important dates', {'fields': ('last_login', 'date_joined')}),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import/', self.admin_site.admin_view(self.import_view), name='accounts_student_import'),
            path('import/template/', self.admin_site.admin_view(self.download_template), name='accounts_student_template'),
            path('add-student/', self.admin_site.admin_view(self.add_student_view), name='accounts_student_add_custom'),
            path('delete-student/', self.admin_site.admin_view(self.delete_student_view), name='accounts_student_delete_custom'),
            path('add-superuser/', self.admin_site.admin_view(self.add_superuser_view), name='accounts_student_add_superuser'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['import_url'] = reverse('admin:accounts_student_import')
        extra_context['add_url'] = reverse('admin:accounts_student_add_custom')
        extra_context['delete_url'] = reverse('admin:accounts_student_delete_custom')
        extra_context['add_superuser_url'] = reverse('admin:accounts_student_add_superuser')
        return super().changelist_view(request, extra_context=extra_context)

    def add_student_view(self, request):
        from .forms import StudentAddForm

        if request.method == 'POST':
            form = StudentAddForm(request.POST)
            if form.is_valid():
                student = form.save(commit=False)
                student.set_unusable_password()
                student.save()
                messages.success(request, f"✅ Talaba qo'shildi: {student.full_name}")
                return HttpResponseRedirect(reverse('admin:accounts_student_changelist'))
        else:
            form = StudentAddForm()

        context = {
            'form': form,
            'title': 'Talaba qo\'shish',
            'opts': self.model._meta,
        }
        return render(request, 'admin/accounts/student/add_student.html', context)

    def delete_student_view(self, request):
        if request.method == 'POST':
            hemis_id = request.POST.get('hemis_id')
            try:
                student = Student.objects.get(hemis_id=hemis_id)
                student.delete()
                messages.success(request, f"🗑️ Talaba o'chirildi: {student.full_name}")
                return HttpResponseRedirect(reverse('admin:accounts_student_changelist'))
            except Student.DoesNotExist:
                messages.error(request, "❌ Talaba topilmadi")
        else:
            messages.error(request, "❌ Noto'g'ri so'rov")

        context = {
            'title': 'Talaba o\'chirish',
            'opts': self.model._meta,
        }
        return render(request, 'admin/accounts/student/delete_student.html', context)

    def add_superuser_view(self, request):
        from .forms import SuperUserAddForm

        if request.method == 'POST':
            form = SuperUserAddForm(request.POST)
            if form.is_valid():
                hemis_id = form.cleaned_data['hemis_id']
                password = form.cleaned_data['password']

                try:
                    student = Student.objects.get(hemis_id=hemis_id)
                    student.is_superuser = True
                    student.is_staff = True
                    student.set_password(password)
                    student.save()
                    messages.success(request, f"✅ SuperUser yangilandi: {student.hemis_id}")
                except Student.DoesNotExist:
                    student = Student.objects.create(
                        hemis_id=hemis_id,
                        full_name='Admin',
                        faculty='Admin',
                        year=1,
                        is_superuser=True,
                        is_staff=True,
                        is_active=True
                    )
                    student.set_password(password)
                    student.save()
                    messages.success(request, f"✅ SuperUser qo'shildi: {student.hemis_id}")

                return HttpResponseRedirect(reverse('admin:accounts_student_changelist'))
        else:
            form = SuperUserAddForm()

        context = {
            'form': form,
            'title': 'SuperUser qo\'shish',
            'opts': self.model._meta,
        }
        return render(request, 'admin/accounts/student/add_superuser.html', context)

    def import_view(self, request):
        if request.method == 'POST':
            file = request.FILES.get('import_file')
            if not file:
                messages.error(request, "Fayl tanlanmadi!")
                return HttpResponseRedirect(request.path)

            filename = file.name.lower()
            try:
                if filename.endswith('.xlsx') or filename.endswith('.xls'):
                    created, updated, errors = self._import_from_excel(file)
                elif filename.endswith('.csv'):
                    created, updated, errors = self._import_from_csv(file)
                else:
                    messages.error(request, "Faqat .xlsx yoki .csv fayl qabul qilinadi!")
                    return HttpResponseRedirect(request.path)

                if created:
                    messages.success(request, f"✅ {created} ta yangi talaba qo'shildi.")
                if updated:
                    messages.warning(request, f"🔄 {updated} ta talaba yangilandi.")
                if errors:
                    for err in errors[:10]:
                        messages.error(request, err)

            except Exception as e:
                messages.error(request, f"Xato yuz berdi: {str(e)}")

            return HttpResponseRedirect(reverse('admin:accounts_student_changelist'))

        context = {
            **self.admin_site.each_context(request),
            'title': "Talabalarni import qilish",
            'template_url': reverse('admin:accounts_student_template'),
            'opts': self.model._meta,
        }
        return render(request, 'admin/accounts/student/import.html', context)

    def _import_from_excel(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        return self._process_rows(list(ws.iter_rows(min_row=2, values_only=True)))

    def _import_from_csv(self, file):
        decoded = TextIOWrapper(file, encoding='utf-8-sig')
        reader = csv.reader(decoded)
        next(reader, None)
        return self._process_rows(list(reader))

    def _process_rows(self, rows):
        created = updated = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            try:
                if not any(row):
                    continue

                hemis_id  = str(row[0]).strip() if row[0] else None
                full_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                birth_date = self._parse_date(row[2]) if len(row) > 2 and row[2] else None
                year      = int(row[3]) if len(row) > 3 and row[3] else 1
                faculty   = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                is_active = str(row[5]).strip().lower() in ('ha', 'true', '1', 'yes') if len(row) > 5 and row[5] else True

                if not hemis_id:
                    errors.append(f"Qator {i}: HEMIS ID bo'sh!")
                    continue

                student, is_new = Student.objects.get_or_create(
                    hemis_id=hemis_id,
                    defaults={
                        'full_name': full_name,
                        'faculty': faculty,
                        'year': year,
                        'birth_date': birth_date,
                        'is_active': is_active,
                    }
                )

                if is_new:
                    student.set_unusable_password()
                    student.save()
                    created += 1
                else:
                    student.full_name = full_name
                    student.faculty = faculty
                    student.year = year
                    student.birth_date = birth_date
                    student.is_active = is_active
                    student.save()
                    updated += 1

            except Exception as e:
                errors.append(f"Qator {i}: {str(e)}")

        return created, updated, errors

    def _parse_date(self, date_str):
        """DD.MM.YYYY yoki YYYY-MM-DD formatidagi sanani parse qilish"""
        from datetime import datetime
        try:
            if isinstance(date_str, datetime):
                return date_str.date()
            date_str = str(date_str).strip()
            
            # DD.MM.YYYY format
            if '.' in date_str:
                parts = date_str.split('.')
                if len(parts) == 3:
                    day, month, year = parts
                    return datetime(int(year), int(month), int(day)).date()
            
            # YYYY-MM-DD format
            elif '-' in date_str:
                parts = date_str.split('-')
                if len(parts) == 3:
                    year, month, day = parts
                    return datetime(int(year), int(month), int(day)).date()
            
            return None
        except:
            return None

    def download_template(self, request):
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Talabalar')

        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center', 'valign': 'vcenter'
        })
        example_fmt = workbook.add_format({'bg_color': '#EBF3FB', 'border': 1})
        note_fmt = workbook.add_format({'italic': True, 'font_color': '#888888'})

        headers = ["HEMIS ID", "To'liq ism", "Tug'ilgan kun (DD.MM.YYYY yoki YYYY-MM-DD)", "Kurs", "Fakultet", "Faol (Ha/Yo'q)"]
        for col, h in enumerate(headers):
            worksheet.write(0, col, h, header_fmt)

        examples = [
            ['2024001', 'Ali Valiyev', '15.06.2000', 2, 'IT', 'Ha'],
            ['2024002', 'Malika Yusupova', '20.03.2001', 1, 'Iqtisodiyot', 'Ha'],
            ['2024003', 'Jasur Toshmatov', '10.09.2002', 3, 'Huquq', "Yo'q"],
        ]
        for row, data in enumerate(examples, 1):
            for col, val in enumerate(data):
                worksheet.write(row, col, val, example_fmt)

        worksheet.write(5, 0, "* 1-qator sarlavha, o'zgartirmang. Faqat ma'lumot qo'shing.", note_fmt)
        worksheet.set_column(0, 0, 12)
        worksheet.set_column(1, 1, 22)
        worksheet.set_column(2, 3, 20)
        worksheet.set_column(4, 4, 8)
        worksheet.set_column(5, 5, 15)
        worksheet.set_row(0, 22)

        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=talabalar_shablon.xlsx'
        return response

    def save_model(self, request, obj, form, change):
        if not change:
            obj.set_unusable_password()
        super().save_model(request, obj, form, change)

    # ↓↓↓ actions_list uchun @unfold_action — bu muhim! ↓↓↓
    @unfold_action(description="Excel ga eksport qilish")
    def export_to_excel(self, request, queryset=None):
        if queryset is None or not queryset.exists():
            queryset = Student.objects.all()

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white'})
        headers = ["HEMIS ID", "To'liq ism", "Tug'ilgan kun", "Kurs", "Fakultet", "Faol", "Ro'yxatga olingan"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, bold)

        for row, student in enumerate(queryset, 1):
            worksheet.write(row, 0, student.hemis_id)
            worksheet.write(row, 1, student.full_name)
            worksheet.write(row, 2, student.birth_date.strftime('%d.%m.%Y') if student.birth_date else '')
            worksheet.write(row, 3, student.year)
            worksheet.write(row, 4, student.faculty)
            worksheet.write(row, 5, "Ha" if student.is_active else "Yo'q")
            worksheet.write(row, 6, student.date_joined.strftime('%d.%m.%Y'))

        worksheet.set_column(0, 0, 15)
        worksheet.set_column(1, 1, 25)
        worksheet.set_column(2, 4, 20)
        worksheet.set_column(5, 6, 15)
        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=talabalar.xlsx'
        return response

    @unfold_action(description="CSV ga eksport qilish")
    def export_to_csv(self, request, queryset=None):
        if queryset is None or not queryset.exists():
            queryset = Student.objects.all()

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename=talabalar.csv'
        writer = csv.writer(response)
        writer.writerow(["HEMIS ID", "To'liq ism", "Tug'ilgan kun", "Kurs", "Fakultet", "Faol", "Ro'yxatga olingan"])

        for student in queryset:
            writer.writerow([
                student.hemis_id, student.full_name,
                student.birth_date.strftime('%d.%m.%Y') if student.birth_date else '',
                student.year, student.faculty,
                "Ha" if student.is_active else "Yo'q",
                student.date_joined.strftime('%d.%m.%Y'),
            ])
        return response